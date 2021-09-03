from django import http
from django.forms.formsets import formset_factory
from django.forms.models import inlineformset_factory
from django.http.response import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponseRedirect,HttpResponse,JsonResponse
from django.contrib.auth.decorators import login_required
from django.urls.base import reverse
from django.views.generic import View, TemplateView, ListView, UpdateView, CreateView, DeleteView
from django.urls import reverse_lazy
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth import login, logout, authenticate
from .forms import * 
from .models import *
from django.contrib import messages
from openpyxl import Workbook
from django.contrib.messages.views import SuccessMessageMixin



class  MixinFormInvalid:
    def form_invalid(self, form):
        response = super().form_invalid(form)
        print("entro")
        print(form)

        if self.request.is_ajax():
            return JsonResponse(form.errors, status=400)
        else:
            return response

def login_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('inicio')
    else:
        if request.method == 'POST':
            form = LoginForm(request.POST)
            if form.is_valid():
                username = form.cleaned_data['username']
                password = form.cleaned_data['password']
                usuario = authenticate(username=username,password=password)
                if usuario is not None and usuario.is_active:
                    login(request,usuario)
                    return HttpResponseRedirect('/inicio/')
        form = LoginForm()
        ctx = {'form': form}
        return render(request, 'inicio/logi.html', ctx)

def logout_view(request):
    logout(request)
    return HttpResponseRedirect('/login/')

@login_required(login_url= '/login/')
def inicio_view(request):
    Inventario = Producto.objects.all()
    ctx = {'Inventario': Inventario }
    return render(request,'inicio/inicio.html',ctx)

def registro(request):
    
    data = {
        'form' : ClientesForm()
    }
    if request.method == 'POST':
        form = ClientesForm(data=request.POST)
        if form.is_valid():
            form.save()
            username = request.POST.get('username',"")
            password = request.POST.get('password',"")
            user=User.objects.create_user(username=username,password=password,last_name="cliente",email=request.POST.get('email',""))
            usuario = authenticate(username=username,password=password)
            login(request, usuario)
            messages.success(request, "Te has registrado correctamente")
            return redirect('/inicio/')
        else:
            print(form)
        data["form"] = form

    return render(request, 'inicio/registrar.html', data)


def perfil(request):
    user=request.user
    cliente=Cliente.objects.get(cedula=user.username)
    user_modifi=Cliente.objects.get(id=user.id)
    data = {
        'form' : ClientesForm(instance=cliente)
    }
    if request.method == 'POST':
        form = ClientesForm(data=request.POST,instance=cliente)
        if form.is_valid():
            form.save()      
            user_modifi.username = form.cleaned_data["cedula"]
            user_modifi.save()
            messages.success(request, "Te has registrado correctamente")
            return redirect('/perfil/')
        else:
            print(form)
        data["form"] = form

    return render(request, 'inicio/perfil.html', data)
# def registrar_view(request):
#     info = "inicializar"
#     if request.method == 'POST':
#         form = PerfilForm(request.POST,request.FILES)
#         if form.is_valid():
#             user = form.save()
#             perfil = Perfil()
#             perfil.usuario = user
#             perfil.telefono = form.cleaned_data['telefono']
#             perfil.direccion = form.cleaned_data['direccion']
#             perfil.cedula= form.cleaned_data['cedula']
#             perfil.save()
#             info = "Guardado Satisfactoriamente"
#             ctx = {'info':info}
#             return render(request, 'inicio/resgistro_exitoso.html',ctx)
#     else:
#         form = PerfilForm()
#         form.fields['username'].help_text = None
#         form.fields['password1'].help_texSuccessMesageMixint = None
#         form.fields['password2'].help_text = None
#     ctx ={'form':form, 'info': info}
#     return render(request, 'inicio/registro_usuario.html',ctx)

# #========================== VISTAS BASADOS EN LA CLASE CATEGORIA =======================#   
class CrearCategoria(CreateView,SuccessMessageMixin):
    model = Categoria
    form_class = CategoriaForm
    template_name='categoria/crear_categoria.html'
    success_url = reverse_lazy('listar_categoria')
    success_message ="Usuario Creado corectamente"    


class ListadoCategoria(ListView):
    model = Categoria
    template_name = 'categoria/listar_categoria.html'
    queryset = Categoria.objects.filter(estado=True)
    context_object_name= 'categorias'

class ActualizarCategoria(UpdateView,SuccessMessageMixin):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'categoria/editar_categoria.html'
    success_url = reverse_lazy('listar_categoria')
    success_message ="Usuario Actualizadoo corectamente"    
    
def eliminar_categoria(request, id):
    categoria = get_object_or_404(Categoria, id=id)
    categoria.delete()
    messages.success(request, "eliminado correctamente!")
    return redirect(to="listar_categoria") 
    
# #========================== VISTAS BASADOS EN LA CLASE CLIENTE =======================#  
class CrearCliente(MixinFormInvalid,CreateView):
    model = Cliente
    template_name='cliente/crear_cliente.html'
    form_class = ClienteForm
    success_url = reverse_lazy('listar_cliente')
    success_message ="Usuario Creado correctamente"    

    def form_valid(self, form):
        self.object = form.save()    
        self.object.save()
        cedula=self.object
     
        user=User.objects.create_user(username=cedula.cedula,password=cedula.cedula,last_name="cliente")

        return http.HttpResponseRedirect(self.get_success_url())

class ListadoCliente(ListView):
    model = Cliente
    template_name = 'cliente/listar_cliente.html'
    form_class = ClienteForm
    queryset = Cliente.objects.filter(estado=True)
    context_object_name= 'clientes'

class ActualizarCliente(MixinFormInvalid,UpdateView,SuccessMessageMixin):
    model = Cliente
    template_name = 'cliente/editar_cliente.html'
    form_class = ClienteUpdateForm
    success_url = reverse_lazy('listar_cliente')
    success_message ="Cliente Actualizadoo corectamente"    

''' class EliminarCliente(DeleteView):
    model = Cliente    
    template_name = 'cliente/cliente_confirm_delete.html'
    def post(self, request, pk, *args, **kwargs):
        object = Cliente.objects.get(id=pk)
        object.estado = False
        object.save()
        return redirect('listar_cliente') '''
        
def eliminar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)
    cliente.delete()
    messages.success(request, "eliminado correctamente!")
    return redirect(to="listar_cliente")

# #========================== VISTAS BASADOS EN LA CLASE COMPRA =======================#  
class CrearCompra(MixinFormInvalid,CreateView,SuccessMessageMixin):
    model = Compra
    form_class = CompraForm
    template_name='compra/crear_compra.html'
    success_url = reverse_lazy('listar_compra')
    success_message ="Compra realizada corectamente"    

class ListadoCompra(ListView):
    model = Compra
    template_name = 'compra/listar_compra.html'
    queryset = Compra.objects.filter(estado=True)
    context_object_name= 'compras'

class ActualizarCompra(MixinFormInvalid,UpdateView,SuccessMessageMixin):
    model = Compra
    form_class = CompraForm
    template_name = 'compra/editar_compra.html'
    success_url = reverse_lazy('listar_compra')
    success_message ="Compra Actualizadoo corectamente"    


# class EliminarCompra(DeleteView):
#     model = Compra   
#     template_name = 'compra/compra_confirm_delete.html'
#     def post(self, request, pk, *args, **kwargs):
#         object = Cliente.objects.get(id=pk)
#         object.estado = False
#         object.save()
#         return redirect('listar_compra')
def eliminar_compra(request, id):
    compra = get_object_or_404(Compra, id=id)
    compra.delete()
    messages.success(request, "eliminado correctamente!")
    return redirect(to="listar_compra")


# #========================== VISTAS BASADOS EN LA CLASE PRODUCTO =======================#  
class CrearProducto(CreateView,SuccessMessageMixin):
    model = Producto
    form_class =ProductoForm
    template_name='producto/crear_producto.html'
    success_url = reverse_lazy('listar_producto')
    success_message ="Usuario Actualizadoo corectamente"    



    def form_invalid(self, form):
        response = super().form_invalid(form)
        print("entro")
        print(form)

        if self.request.is_ajax():
            return JsonResponse(form.errors, status=400)
        else:
            return response

    

class ListadoProducto(ListView):
    model = Producto
    template_name = 'producto/listar_producto.html'
    queryset = Producto.objects.filter(estado=True)
    context_object_name= 'productos'

class ActualizarProducto(MixinFormInvalid,UpdateView,SuccessMessageMixin):
    model = Producto
    form_class = ProductoForm
    template_name = 'producto/editar_producto.html'
    success_url = reverse_lazy('listar_producto')
    success_message ="Usuario Actualizadoo corectamente"    

        
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.delete()
    messages.success(request, "eliminado correctamente!")
    return redirect(to="listar_producto")

# #========================== VISTAS BASADOS EN LA CLASE PROVEEDOR =======================#  
class CrearProveedor(MixinFormInvalid,CreateView,SuccessMessageMixin):
    model = Proveedor
    form_class =ProveedorForm
    template_name='proveedor/crear_proveedor.html'
    success_url = reverse_lazy('listar_proveedor')
    success_message ="Usuario creado corectamente"    

class ListadoProveedor(ListView):
    model = Proveedor
    template_name = 'proveedor/listar_proveedor.html'
    queryset = Proveedor.objects.filter(estado=True)
    context_object_name= 'proveedores'

class ActualizarProveedor(MixinFormInvalid,UpdateView,SuccessMessageMixin):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'proveedor/editar_proveedor.html'
    success_url = reverse_lazy('listar_proveedor')
    success_message ="Proveedor Actualizadoo corectamente"    


''' class EliminarProveedor(DeleteView):
    model = Proveedor   
    template_name = 'proveedor/proveedor_confirm_delete.html'
    def post(self, request, pk, *args, **kwargs):
        object = Proveedor.objects.get(id=pk)
        object.estado = False
        object.save()
        return redirect('listar_proveedor')
 '''
def eliminar_proveedor(request, id):
        proveedor = get_object_or_404(Proveedor, id=id)
        proveedor.delete()
        messages.success(request, "eliminado correctamente!")
        return redirect(to="listar_proveedor")
# #========================== VISTAS BASADOS EN LA CLASE VENTA =======================#  
import json
def ajax_load_project(request):
    print("askjfdn")
    search_qs = Producto.objects.filter(nombre__startswith=request.GET['search'])
    results = []
    for r in search_qs:
        results.append(r.nombre)
    resp =  json.dumps(results)
    return HttpResponse(resp, content_type='application/json')

class CrearVenta(MixinFormInvalid,CreateView,SuccessMessageMixin):
   
    model = Venta
    form_class =VentaForm
    template_name='venta/crear_venta.html'
    success_url = reverse_lazy('listar_venta')
    success_message ="Usuario Actualizadoo corectamente"    

    
        
    
   

class ListadoVenta(ListView):
    model = Venta
    template_name = 'venta/listar_venta.html'
    queryset = Venta.objects.all()
    context_object_name= 'ventas'

class ActualizarVenta(MixinFormInvalid,UpdateView, SuccessMessageMixin):
    model = Venta
    form_class = VentaForm
    template_name = 'venta/editar_venta.html'
    success_url = reverse_lazy('listar_venta')
    success_message ="Usuario Actualizadoo corectamente"    


''' class EliminarVenta(DeleteView):
    model = Venta   
    template_name = 'venta/venta_confirm_delete.html'
    def post(self, request, pk, *args, **kwargs):
        object = Venta.objects.get(id=pk)
        object.estado = False
        object.save()
        return redirect('listar_venta') '''

def eliminar_venta(request, id):
    venta = get_object_or_404(Venta, id=id)
    venta.delete()
    messages.success(request, "eliminado correctamente!")
    return redirect(to="listar_venta")
# #========================== VISTAS BASADOS EN LA CLASE ReporteVENTA =======================#  
""" class ReporteVenta(TemplateView):
    def get(self, request, pk, *args, **kwargs):
       _venta= Venta.objects.all()
        wb = Workbook()  """

class ReporteproductoExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        _producto = Producto.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE PRODUCTOS'
        ws.merge_cells('B1:E1')
        ws['B3'] = 'Codigo'
        ws['C3'] = 'nombre'
        ws['D3'] = 'Descripcion'
        ws['E3'] = 'Existencia'
        ws['F3'] = 'Foto'
        ws['G3'] = 'Precio'
        ws['H3'] = 'Categoria'
        ws['I3'] = 'Proveedor'
        
        cont = 4
        for productos in _producto:
            ws.cell(row = cont, column = 2).value = productos.producto.codigo
            ws.cell(row = cont, column = 3).value = productos.producto.nombre
            ws.cell(row = cont, column = 4).value = productos.producto.descripcion
            ws.cell(row = cont, column = 5).value = productos.producto.existencia
            ws.cell(row = cont, column = 6).value = productos.producto.foto_producto1
            ws.cell(row = cont, column = 7).value = productos.categoria.fk_categoria
            ws.cell(row = cont, column = 8).value = productos.proveedor.fk_proveedor
            cont+=1
        nombre_archivo = "ReporProductoExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        # wb.save(response)
        return response